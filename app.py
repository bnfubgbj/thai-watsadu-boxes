import streamlit as st
import fitz
import re
import io
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string as ci
from datetime import date
import os

st.set_page_config(page_title="ไทวัสดุ — คำนวณกล่องสินค้า", page_icon="📦", layout="wide")

st.markdown("""
<style>
div[data-testid="stSidebar"] { background: #1a1a2e; }
div[data-testid="stSidebar"] * { color: #eee !important; }
.brand { background: linear-gradient(135deg,#c62828,#e53935); color:#fff; padding:1.2rem 1.5rem;
         border-radius:12px; margin-bottom:1.5rem; }
.brand h1 { margin:0; font-size:1.6rem; }
.brand p  { margin:0; opacity:.85; font-size:.9rem; }
.stat-box { background:#fff; border:1px solid #e0e0e0; border-radius:10px;
            padding:1rem; text-align:center; }
.stat-box .num { font-size:2rem; font-weight:700; color:#c62828; }
.stat-box .lbl { font-size:.8rem; color:#888; }
.branch-card { background:#fff; border:1px solid #e8e8e8; border-radius:10px;
               padding:1rem 1.2rem; margin-bottom:.6rem; }
.badge { display:inline-block; padding:2px 8px; border-radius:20px;
         font-size:.75rem; font-weight:600; margin:1px; }
.b-mixed  { background:#fff8e1; color:#f57f17; border:1px solid #ffe082; }
.b-foam   { background:#e8f5e9; color:#2e7d32; border:1px solid #a5d6a7; }
.b-canvas { background:#ffebee; color:#c62828; border:1px solid #ef9a9a; }
.b-num    { background:#e3f2fd; color:#1565c0; border:1px solid #90caf9; }
.ok   { color:#2e7d32; font-weight:600; }
.warn { color:#c62828; font-weight:600; }
</style>
""", unsafe_allow_html=True)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "branches.xlsx")

@st.cache_data
def load_branch_master():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Store รหัสสาขา"]
    data = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[2] and str(row[2]).strip().isdigit():
            upfront = int(row[2])
            data[upfront] = {
                "name_en": (row[3] or "").strip(),
                "name_th": (row[8] or "").strip(),
            }
    return data

BRANCH_MASTER = load_branch_master()

# ─── Box calculation ───────────────────────────────────────────────────────────
def calc_boxes(canvas_pairs, foam_dozen) -> list[dict]:
    boxes, cp, fd = [], int(canvas_pairs), float(foam_dozen)
    # mixed: ฟองน้ำ 1 โหล + ผ้าใบ 6 คู่
    mixed = min(int(fd), int(cp) // 6)
    for _ in range(mixed):
        boxes.append({"type": "mixed", "label": "ฟองน้ำ 1 โหล + ผ้าใบ 6 คู่"})
        fd -= 1; cp -= 6
    # ฟองน้ำล้วน 2 โหล/กล่อง
    while fd >= 2:
        boxes.append({"type": "foam", "label": "ฟองน้ำ 2 โหล"})
        fd -= 2
    # เศษฟองน้ำ 1 โหล
    if fd >= 1:
        if cp >= 6:
            # ผ้าใบเหลือ >= 6 → mixed
            boxes.append({"type": "mixed", "label": "ฟองน้ำ 1 โหล + ผ้าใบ 6 คู่"})
            cp -= 6
        elif cp > 0:
            # ผ้าใบเหลือ < 6 แต่มีอยู่ → รวมกล่องเดียว
            boxes.append({"type": "mixed", "label": f"ฟองน้ำ 1 โหล + ผ้าใบ {cp} คู่"})
            cp = 0
        else:
            # ไม่มีผ้าใบเลย
            boxes.append({"type": "foam", "label": "ฟองน้ำ 1 โหล"})
        fd = 0
    # ผ้าใบที่เหลือ
    while cp >= 12:
        boxes.append({"type": "canvas", "label": "ผ้าใบ 12 คู่"})
        cp -= 12
    if cp > 0:
        boxes.append({"type": "canvas", "label": f"ผ้าใบ {cp} คู่"})
    return boxes

# ─── PDF Parser ────────────────────────────────────────────────────────────────
def parse_pdf(pdf_bytes: bytes, filename: str) -> dict:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    result = {"so_no": "", "po_no": "", "invoice_no": "", "ship_date": "", "po_items": {"canvas": 0, "foam200": 0, "foam212": 0},
              "branches": {}, "errors": [], "filename": filename}

    first_text = doc[0].get_text()
    m_so   = re.search(r'SO\d+-\d+', first_text)
    m_inv  = re.search(r'(?:ใบสสงซซอเลขททส|ใบสั่งซื้อเลขที่)\s+(\d+)', first_text)
    m_ship = re.search(r'Ship Date\s*([\d/]+)', first_text)
    if m_so:   result["so_no"]       = m_so.group(0)
    else:
        # หา SO จากชื่อไฟล์ เช่น SO6906-0067
        m_so_fn = re.search(r'SO\d+-\d+', filename)
        if m_so_fn: result["so_no"] = m_so_fn.group(0)
    if m_inv:  result["po_no"]      = m_inv.group(1)
    if m_ship: result["ship_date"]  = m_ship.group(1)

    # รูปแบบใหม่ (CRC Thai Watsadu Ltd.) — Ship Date คือบรรทัดวันที่ 2 ติดกัน, PO ถัดไป
    if not result["po_no"] or not result["ship_date"]:
        lines_ft = first_text.split('\n')
        DAY_PAT = r'^(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun),'
        for idx_ft, ln in enumerate(lines_ft):
            if re.match(DAY_PAT, ln.strip()) and idx_ft > 0 and re.match(DAY_PAT, lines_ft[idx_ft-1].strip()):
                if not result["ship_date"]:
                    result["ship_date"] = ln.strip()
                if idx_ft + 1 < len(lines_ft):
                    next_ln = lines_ft[idx_ft + 1].strip()
                    m_po = re.match(r'^(\d{10})', next_ln)
                    if m_po and not result["po_no"]:
                        result["po_no"] = m_po.group(1)
                break

    po_pages = []
    dist_pages = []

    for i in range(len(doc)):
        text = doc[i].get_text()
        if "ใบแบงสสนคคา" in text or "ใบแบบงสสนคคา" in text or "ใบแบ่งสินค้า" in text or "กระจายไปสาขา" in text or "กระจายไปยังสาขา" in text:
            dist_pages.append(i)
        elif "ใบสสงซซอ" in text or "จสานวนสสง" in text:
            po_pages.append(i)

    for pi in po_pages:
        text = doc[pi].get_text()
        lines = text.split('\n')
        for j, line in enumerate(lines):
            if 'รองเทคาผคาใบ' in line or 'NANYANG 205' in line:
                ctx = '\n'.join(lines[j:j+15])
                for m2 in re.finditer(r'(\d+\.\d+)\s*\n?\s*EACH', ctx):
                    qty = float(m2.group(1))
                    if qty >= 6:
                        result["po_items"]["canvas"] += qty
                    break
            elif 'รองเทคาแตะ' in line or 'หหนทบ' in line or 'สวม' in line:
                ctx = '\n'.join(lines[j:j+15])
                is_212 = '212' in '\n'.join(lines[max(0,j-2):j+5]) or 'สวม' in line
                is_200 = '200' in '\n'.join(lines[max(0,j-2):j+5]) or 'หหนทบ' in line
                for m2 in re.finditer(r'(\d+\.\d+)\s*\n?\s*EACH', ctx):
                    qty = float(m2.group(1))
                    if qty >= 12:
                        if is_212:
                            result["po_items"]["foam212"] += qty
                        elif is_200:
                            result["po_items"]["foam200"] += qty
                    break

    current_branch = None
    for pi in dist_pages:
        text = doc[pi].get_text()
        lines = text.split('\n')

        idx = 0
        while idx < len(lines):
            line = lines[idx].strip()

            bm = re.search(r'กระจายไป(?:สาขา|ยังสาขา)\s+(\d{5})\s+(.*)', line)
            if bm:
                code = int(bm.group(1))
                name_raw = bm.group(2).strip()
                current_branch = code
                if code not in result["branches"]:
                    master = BRANCH_MASTER.get(code, {})
                    result["branches"][code] = {
                        "name": master.get("name_en") or name_raw,
                        "name_th": master.get("name_th") or "",
                        "canvas": 0, "foam200": 0, "foam212": 0,
                    }
                idx += 1
                continue

            if current_branch and ('รองเทคาผคาใบ' in line or 'รองเทคาแตะ' in line
                                      or 'รองเท้าผ้าใบ' in line or 'รองเท้าแตะ' in line):
                product_line = line
                qty = 0
                for k in range(idx + 1, min(idx + 10, len(lines))):
                    m = re.match(r'^\s*([\d]+\.[\d]+)\s*$', lines[k])
                    if m:
                        qty = float(m.group(1))
                        break
                if qty > 0:
                    if 'ผคาใบ' in product_line or 'ผ้าใบ' in product_line or '205' in product_line:
                        result["branches"][current_branch]["canvas"] += qty
                    elif '212' in product_line or ('สวม' in product_line and '200' not in product_line):
                        result["branches"][current_branch]["foam212"] += qty
                    elif '200' in product_line or 'หหนทบ' in product_line or 'หูหนีบ' in product_line or 'แตะ' in product_line:
                        result["branches"][current_branch]["foam200"] += qty

            idx += 1
    doc.close()
    return result


# ─── Excel generation ──────────────────────────────────────────────────────────
def generate_excel(all_branches: list[dict], company: str, invoice_no: str,
                   invoice_date, total_boxes: int) -> bytes:
    wb_tpl = load_workbook(TEMPLATE_PATH)
    tpl_ws = wb_tpl["ใบปะหน้า"]
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Sheet สรุป — แก้ "Invoice No." → "PO No."
    ws_sum = wb_out.create_sheet("สรุปทุกสาขา")
    headers = ["ลำดับ", "SO No.", "PO No.", "สาขา (EN)", "สาขา (TH)", "รหัส",
               "ผ้าใบ (คู่)", "ฟองน้ำ200 (คู่)", "ฟองน้ำ212 (คู่)", "กล่อง", "กล่องที่"]
    ws_sum.append(headers)
    for cell in ws_sum[1]:
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C62828")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, b in enumerate(all_branches):
        bx_count = len(b["boxes"])
        box_nums = ", ".join(f"{k+1}/{bx_count}" for k in range(bx_count))
        ws_sum.append([i+1, b.get("so_no",""), b.get("po_no",""), b["name"], b["name_th"], str(b["upfront"]),
                       b["canvas"], b["foam200"], b["foam212"],
                       bx_count, box_nums])

    for col in ws_sum.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_sum.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws_sum.freeze_panes = "A2"

    # Sheet ใบปะหน้าทั้งหมด (2 ใบ/หน้า แนวตั้ง)
    from openpyxl.styles import Border, Side
    from openpyxl.worksheet.pagebreak import Break
    thin = Side(style='thin')
    ROWS_PER_LABEL = 9
    LABELS_PER_PAGE = 2

    ws_all = wb_out.create_sheet("ใบปะหน้าทั้งหมด")

    for col_l, cd in tpl_ws.column_dimensions.items():
        ws_all.column_dimensions[col_l].width = cd.width

    all_labels = []
    for b in all_branches:
        bx_total = len(b["boxes"])
        for box_idx, box in enumerate(b["boxes"]):
            all_labels.append({"b": b, "box_idx": box_idx, "bx_total": bx_total, "box": box})

    for label_idx, item in enumerate(all_labels):
        b = item["b"]
        box_idx = item["box_idx"]
        bx_total = item["bx_total"]
        box_label = f"{box_idx+1}/{bx_total}"
        ro = label_idx * ROWS_PER_LABEL  # row offset

        for merged in tpl_ws.merged_cells.ranges:
            ws_all.merge_cells(
                start_row=merged.min_row + ro, start_column=merged.min_col,
                end_row=merged.max_row + ro,   end_column=merged.max_col,
            )

        for row in tpl_ws.iter_rows(min_row=1, max_row=ROWS_PER_LABEL):
            r = row[0].row
            if r in tpl_ws.row_dimensions:
                ws_all.row_dimensions[r + ro].height = tpl_ws.row_dimensions[r].height
            for cell in row:
                nc = ws_all.cell(row=cell.row + ro, column=cell.column)
                if cell.has_style:
                    nc.font      = copy.copy(cell.font)
                    nc.border    = copy.copy(cell.border)
                    nc.fill      = copy.copy(cell.fill)
                    nc.alignment = copy.copy(cell.alignment)
                if cell.value and not str(cell.value).startswith("=VLOOKUP"):
                    nc.value = cell.value

        for img in tpl_ws._images:
            try:
                img_copy = copy.deepcopy(img)
                try:
                    img_copy.anchor._from.row += ro
                    if hasattr(img_copy.anchor, 'to') and img_copy.anchor.to:
                        img_copy.anchor.to.row += ro
                except Exception:
                    pass
                ws_all.add_image(img_copy)
            except Exception:
                pass

            ws_all.cell(row=3+ro, column=ci("F")).value = b["upfront"] if str(b["upfront"]).isdigit() else ""
        ws_all.cell(row=3+ro, column=ci("D")).value = b["name"]
        ws_all.cell(row=4+ro, column=ci("E")).value = b["name_th"] or b["name"]
        ws_all.cell(row=5+ro, column=ci("D")).value = f"   {company}"
        ws_all.cell(row=6+ro, column=ci("E")).value = ""
        ws_all.cell(row=7+ro, column=ci("E")).value = b.get("ship_date","") or str(invoice_date)

        cell_a9 = ws_all.cell(row=9+ro, column=1)
        cell_a9.value = f"กล่องที่  {box_label}        รวม  {bx_total}  กล่อง"
        cell_a9.font = Font(name="AngsanaUPC", size=25, bold=True)
        cell_a9.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, 13):
            nc = ws_all.cell(row=9+ro, column=col_idx)
            nc.border = Border(
                top=nc.border.top, left=nc.border.left,
                right=nc.border.right, bottom=thin,
            )

        if (label_idx + 1) % LABELS_PER_PAGE == 0 and label_idx < len(all_labels) - 1:
            ws_all.row_breaks.append(Break(id=(label_idx + 1) * ROWS_PER_LABEL))

    total_rows = len(all_labels) * ROWS_PER_LABEL
    ws_all.print_area = f"A1:L{total_rows}"
    ws_all.page_setup.paperSize   = ws_all.PAPERSIZE_A4
    ws_all.page_setup.orientation = ws_all.ORIENTATION_PORTRAIT
    ws_all.page_setup.fitToPage   = True
    ws_all.page_setup.fitToWidth  = 1
    ws_all.page_setup.fitToHeight = 0
    ws_all.page_margins.left   = 0.3
    ws_all.page_margins.right  = 0.3
    ws_all.page_margins.top    = 0.3
    ws_all.page_margins.bottom = 0.3
    ws_all.page_margins.header = 0
    ws_all.page_margins.footer = 0
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.read()

# ─── Session state ─────────────────────────────────────────────────────────────
if "parsed_files" not in st.session_state:
    st.session_state.parsed_files = {}
if "all_branches" not in st.session_state:
    st.session_state.all_branches = []

# ─── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ ตั้งค่า")
    st.markdown("---")
    st.markdown("### 📋 เงื่อนไขกล่อง")
    st.markdown("""
- 👟 ผ้าใบ **12 คู่** = 1 กล่อง
- 🩴 ฟองน้ำ **2 โหล** = 1 กล่อง
- 🔀 ฟองน้ำ **1 โหล** + ผ้าใบ **6 คู่** = 1 กล่อง
- 🔀 ฟองน้ำ **1 โหล** + ผ้าใบ **1–5 คู่** = 1 กล่อง (เศษที่เหลือ)
    """)

company = "นันยางมาร์เก็ตติ้ง จำกัด"
invoice_no = ""
invoice_date = date.today()

# ─── Main ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="brand">
  <h1>📦 ไทวัสดุ — คำนวณกล่องสินค้า</h1>
  <p>อัปโหลดใบแบ่งสินค้า PDF · คำนวณกล่องอัตโนมัติ · ดาวน์โหลด Excel แยกสาขา · ฟรี 100%</p>
</div>
""", unsafe_allow_html=True)

st.subheader("📂 อัปโหลดใบแบ่งสินค้า (PDF)")
uploaded = st.file_uploader("ลากไฟล์มาวางที่นี่ หรือคลิกเพื่อเลือก",
                             type=["pdf"], accept_multiple_files=True,
                             label_visibility="collapsed")

col1, col2 = st.columns([2, 8])
with col1:
    analyze_btn = st.button("🔍 วิเคราะห์", type="primary", disabled=not uploaded)
with col2:
    if st.button("🗑️ ล้างทั้งหมด"):
        st.session_state.parsed_files = {}
        st.session_state.all_branches = []
        st.rerun()

# ─── Analyse ───────────────────────────────────────────────────────────────────
if analyze_btn and uploaded:
    st.session_state.parsed_files = {}
    st.session_state.all_branches = []
    merged_map = {}

    prog = st.progress(0, text="กำลังอ่าน PDF...")
    for i, f in enumerate(uploaded):
        prog.progress((i) / len(uploaded), text=f"กำลังอ่าน: {f.name}")
        parsed = parse_pdf(f.read(), f.name)
        st.session_state.parsed_files[f.name] = parsed

        for code, b in parsed["branches"].items():
            key = f"{parsed['po_no']}_{code}"
            if key in merged_map:
                merged_map[key]["canvas"] += b["canvas"]
                merged_map[key]["foam200"] += b["foam200"]
                merged_map[key]["foam212"] += b["foam212"]
            else:
                merged_map[key] = {**b, "upfront": code, "so_no": parsed.get("so_no",""), "po_no": parsed["po_no"],
                                   "invoice_no": parsed.get("invoice_no",""),
                                   "ship_date": parsed.get("ship_date",""),
                                   "srcFiles": [f.name]}
        prog.progress((i+1) / len(uploaded))

    for b in merged_map.values():
        foam_total = (b["foam200"] + b["foam212"]) / 12
        b["boxes"] = calc_boxes(int(b["canvas"]), foam_total)

    st.session_state.all_branches = list(merged_map.values())
    prog.empty()


# ─── Results ───────────────────────────────────────────────────────────────────
branches = st.session_state.all_branches
if branches:
    total_boxes = sum(len(b["boxes"]) for b in branches)
    total_pairs = sum(b["canvas"] + b["foam200"] + b["foam212"] for b in branches)

    st.markdown("---")
    st.subheader("📊 สรุปผลการคำนวณ")
    c1, c2, c3, c4 = st.columns(4)
    for col, num, lbl in [
        (c1, len(uploaded) if uploaded else "-", "ไฟล์"),
        (c2, len(branches), "สาขา"),
        (c3, total_boxes, "กล่องทั้งหมด"),
        (c4, f"{total_pairs:,.0f}", "คู่ทั้งหมด"),
    ]:
        col.markdown(f'<div class="stat-box"><div class="num">{num}</div><div class="lbl">{lbl}</div></div>',
                     unsafe_allow_html=True)

    st.markdown("---")
    st.subheader("🏪 สรุปรายละเอียดแต่ละสาขา")

    import pandas as pd
    rows = []
    for b in branches:
        bx_count = len(b["boxes"])
        box_labels = ", ".join(f"{k+1}/{bx_count}" for k in range(bx_count))
        box_types = " | ".join(bx["label"] for bx in b["boxes"])
        rows.append({
            "SO No.":      b.get("so_no",""),
            "PO No.":      b.get("po_no",""),
            "สาขา (EN)":  b["name"],
            "สาขา (TH)":  b["name_th"],
            "รหัส":        str(b["upfront"]),
            "ผ้าใบ (คู่)": int(b["canvas"]),
            "F200 (คู่)":  int(b["foam200"]),
            "F212 (คู่)":  int(b["foam212"]),
            "กล่อง":       bx_count,
            "กล่องที่":    box_labels,
            "รายละเอียดกล่อง": box_types,
        })
    df = pd.DataFrame(rows)
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "SO No.":      st.column_config.TextColumn("SO No.", width="small"),
            "PO No.":      st.column_config.TextColumn("PO No.", width="small"),
            "สาขา (EN)":  st.column_config.TextColumn("สาขา (EN)", width="medium"),
            "สาขา (TH)":  st.column_config.TextColumn("สาขา (TH)", width="medium"),
            "รหัส":        st.column_config.TextColumn("รหัส", width="small"),
            "ผ้าใบ (คู่)": st.column_config.NumberColumn("ผ้าใบ (คู่)", width="small"),
            "F200 (คู่)":  st.column_config.NumberColumn("F200 (คู่)", width="small"),
            "F212 (คู่)":  st.column_config.NumberColumn("F212 (คู่)", width="small"),
            "กล่อง":       st.column_config.NumberColumn("กล่อง", width="small"),
            "กล่องที่":    st.column_config.TextColumn("กล่องที่", width="medium"),
            "รายละเอียดกล่อง": st.column_config.TextColumn("รายละเอียด", width="large"),
        }
    )

    st.markdown("---")
    st.subheader("📥 ดาวน์โหลด Excel")
    col_dl, col_info = st.columns([2, 4])
    with col_dl:
        if st.button("⚙️ สร้าง Excel", type="primary"):
            with st.spinner("กำลังสร้างไฟล์..."):
                try:
                    xlsx = generate_excel(branches, company, invoice_no, invoice_date, total_boxes)
                    st.download_button(
                        "📥 ดาวน์โหลด Excel",
                        data=xlsx,
                        file_name=f"ไทวัสดุ_ใบรับสินค้า_{invoice_no or 'export'}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"เกิดข้อผิดพลาด: {e}")
    with col_info:
        st.info(f"จะสร้าง **{total_boxes + 1} sheet** — สรุปทุกสาขา + ใบรับสินค้า {total_boxes} ใบ")
